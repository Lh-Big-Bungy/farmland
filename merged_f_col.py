from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
import glob
def get_megerd_data():
    # 查找当前目录下所有符合模式的文件
    files = glob.glob("*（*）.xlsx")

    # 如果找到了至少一个文件
    if files:
        wb = load_workbook(files[0])  # 加载第一个匹配的文件
        print(f"加载的文件为: {files[0]}")
    else:
        print("未找到匹配的文件")
    # 加载 Excel 文件
    ws = wb.active

    # 获取合并单元格范围
    merged_ranges = ws.merged_cells.ranges

    merged_blocks = []  # 用于存储每块合并区域的行号列表

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        # 如果是 D 列（第 4 列）
        if min_col == 4 and max_col == 4:
            # 生成这一块合并区域的所有行号，并加到列表中
            merged_block = list(range(min_row, max_row + 1))
            merged_blocks.append(merged_block)

    print("D列中被合并的区域行号块为：", merged_blocks)

    # 打印被合并的行
    merged_dict = {}
    for block in merged_blocks:
        name = ws.cell(row=block[0], column=2).value  # 取合并块首行的“名字”
        area_list = []
        for row in block:
            area = ws.cell(row=row, column=5).value  # 面积是第5列
            area_list.append(area)

        if name in merged_dict:
            merged_dict[name].append(area_list)  # 添加一个新的区域
        else:
            merged_dict[name] = [area_list]  # 创建新的区域列表
    print(merged_dict)
    return merged_dict

def merged_f_col(merged_dict):
# 读取 Excel 文件
    file_path = "output_file.xlsx"  # 你的 Excel 文件
    wb = load_workbook(file_path)
    # 获取所有 sheet 名称，并排除最后一个
    sheets = wb.sheetnames

    for sheet in sheets:
        if sheet in merged_dict:
            col_area_dict = {}
            ws = wb[sheet]
            c_col = ws["C"]
            # print(f"\n📄 Sheet: {sheet} - C列所有值:")
            start_end_list = []
            for i in merged_dict[sheet]:
                temp_se_list = []
                for j in i:
                    start_index = False
                    end_index = False
                    s_e_list = []
                    for cell in c_col[3:-1]:
                        row = cell.row
                #        print(cell.coordinate, "->", cell.value)
                        if cell.value == j and not start_index:  #  获取开始的行数
                            start_index = row
                        elif cell.value != j and start_index and not end_index:
                                end_index = row - 1
                    s_e_list.append(start_index)
                    s_e_list.append(end_index)
                    print(s_e_list)
                    temp_se_list.append(s_e_list)
                start_end_list.append(temp_se_list)  # 为了分区，出来类似这样的列表：[[[4, 7], [8, 10]], [[11, 14], [15, 17]]]
            print(6666, start_end_list)
            merged_list = []
            for x in range(len(start_end_list)):
                for y in range(len(start_end_list[x]) - 1):
                    if (start_end_list[x][y][1] + 1) == start_end_list[x][y+1][0]:  # 行号尾首相连表明是挨着的，是同一块土地
                        if start_end_list[x][y+1][0] not in merged_list:
                            merged_list.append(start_end_list[x][y][0])
                            merged_list.append(start_end_list[x][y+1][1])
                            print(77777777, merged_list)
                ws.merge_cells(f"F{merged_list[0]}:F{merged_list[-1]}")
                merged_list = []


            wb.save(file_path)

def f_col_merged_run():
    merged_dict = get_megerd_data()

    merged_f_col(merged_dict)
if __name__ == '__main__':
    f_col_merged_run()