from datetime import datetime
from openpyxl.styles import Alignment
import os
from openpyxl import load_workbook, Workbook
import cn2an
from hanziconv import HanziConv  # 简体转繁体
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP

def round_half_up(value, digits):
    return float(Decimal(str(value)).quantize(Decimal('1.' + '0' * digits), rounding=ROUND_HALF_UP))

def header_into_excel(name, village_name, date, excel_header):
    # **Step 1: 定义 Excel 文件名**
    file_name = 'output_file.xlsx'

    # **Step 2: 获取当前应该使用的 sheet 编号**
    if os.path.exists(file_name):
        flag = True
        sheet_name = name
    else:
        flag = False
        sheet_name = name
        wb = Workbook()  # 创建新的 Excel
        ws = wb.active  # 获取默认的 Sheet
        ws.title = sheet_name  # 修改默认 Sheet 的名称为传入的名称
        wb.save(file_name)  # 先保存文件

    # **Step 3: 打开 Excel，准备操作**
    wb = load_workbook(file_name)

    # 兼容名字不排在一起或者村集体也有地上附着物的情况
    if name in wb.sheetnames and len(wb.sheetnames) != 1:  # 避免name为第一个时就进入这个处理
        return name

    # **Step 4: 创建新的 Sheet**

    if flag:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb.active

    # 获取所有 Sheet 的名称列表
    sheet_names = wb.sheetnames
    # 判断长度并插入换行符
    if len(excel_header) > 27:
        # 这里按汉字长度截取，假设汉字为全角
        first_part = excel_header[:27]
        second_part = excel_header[27:]
        excel_header = first_part + '\n' + second_part
        char_per_line = 27
        lines = len(excel_header) // char_per_line + 1
        ws.row_dimensions[1].height = 14 * lines  # 基础行高14
    # 获取 Sheet 的个数
    sheet_count = len(sheet_names)
    # **Step 5: 合并 A1 到 D1，并填入表头**
    merge_range = "A1:F1"
    ws.merge_cells(merge_range)
    ws["A1"].value = excel_header
    ws["A1"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # 单独设置A1
    merge_range2 = "A2:F2"
    ws.merge_cells(merge_range2)
    ws["A2"].value = f"户号:0000{sheet_count:<6}  户主:{name:<12}  住址:{village_name:<10}  {date}"
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    # **Step 6: 合并 "户号" 这一列的单元格（如 A2:B2）**
    # ws["A2"].value = f"户号:0000{sheet_count}"
    ws["A3"].value = f"项    目"
    ws.merge_cells("B2:C2")  # 让第一行数据的"户主"占两格
    # ws["B2"].value = f"户主:{name}"
    ws["B3"].value = f"单位"
    # ws["D2"].value = f"住址:{village_name}"
    ws["C3"].value = f"数量"
    # ws.merge_cells("E2:F2")  # 让第一行数据的"时间"占两格
    # ws["E2"].value = f"{date}"
    ws["D3"].value = f"单价(元)"
    ws["E3"].value = f"小计(元)"
    ws["F3"].value = f"备注"
    cell_list = ["A3", "B3", "C3", "D3", "E3", "F3"]
    # cell_list2 = ["B2", "D2"]
    for i in cell_list:
        ws[i].alignment = Alignment(horizontal="center", vertical="center")
    # for j in cell_list2:
    #     ws[j].alignment = Alignment(horizontal="left", vertical="center")

    wb.save(file_name)
    return sheet_name

def handle_handi(sheet_name, area, year, roman, yingxiang, buchang, lingxing, yingxiangdanjia, buchangdanjia):
    """处理地上附着物非户主所有的情况"""
    new_data = [
        [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
        [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
        ['耕地零星林木', '亩', area, 2000.00, lingxing],
    ]
    data_into_excel(sheet_name, new_data)



def handle_lindi(sheet_name, area, year, roman, yingxiang, buchang, yingxiangdanjia, buchangdanjia):
    """处理林地、建设用地、道路、沟渠的补偿数据"""
    new_data = [
        [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
        [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
    ]
    data_into_excel(sheet_name, new_data)

def handle_beifen(sheet_name, number, beifen):
    """处理有主碑坟的补偿数据"""
    new_data = [
        ['有主碑坟', '座', float(number), 5000.00, beifen],
    ]
    data_into_excel(sheet_name, new_data)

def handle_pufen(sheet_name, number, pufen):
    """处理有主普坟的补偿数据"""
    new_data = [
        ['有主普坟', '座', float(number), 3000.00, pufen],
    ]
    data_into_excel(sheet_name, new_data)

def handle_shaichang(data_list, sheet_name, area, year, roman, yingxiang, buchang, yingxiangdanjia, buchangdanjia, shaichang):
    if '地上归' in data_list[1]:
        """处理地上附着物非户主所有的情况"""
        new_data = [
            [f'临时用地{year}年补偿', 'm2', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', 'm2', area, yingxiangdanjia, yingxiang],
        ]
        data_into_excel(sheet_name, new_data)

        new_data2 = [
            ['晒场硬化', 'm2', area, 40.00, shaichang],
        ]
        name = data_list[1].split('归')[1].split('）')[0]
        return new_data2, name
    else:
        """处理晒场硬化的补偿数据"""
        new_data = [
            [f'临时用地{year}年补偿', 'm2', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', 'm2', area, yingxiangdanjia, yingxiang],
            ['晒场硬化', 'm2', area, 40.00, shaichang],
        ]
        data_into_excel(sheet_name, new_data)
        return None, None  # 确保在所有情况下都有返回值

def handle_shucaidapeng(data_list, sheet_name, area, year, roman, yingxiang, buchang, yingxiangdanjia, buchangdanjia, shucaidapeng):
    if '地上归' in data_list[1]:
        """处理地上附着物非户主所有的情况"""
        new_data = [
            [f'临时用地{year}年补偿', 'm2', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', 'm2', area, yingxiangdanjia, yingxiang],
        ]
        data_into_excel(sheet_name, new_data)

        new_data2 = [
            ['蔬菜大棚拆迁', 'm2', area, 45.00, shucaidapeng],
        ]
        name = data_list[1].split('归')[1].split('）')[0]
        return new_data2, name
    else:
        """处理蔬菜大棚的补偿数据"""
        new_data = [
            [f'临时用地{year}年补偿', 'm2', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', 'm2', area, yingxiangdanjia, yingxiang],
            ['蔬菜大棚拆迁', 'm2', area, 45.00, shucaidapeng],
        ]
        data_into_excel(sheet_name, new_data)
        return None, None  # 确保在所有情况下都有返回值

def handle_shuijing(sheet_name, number, shuijing):
    """处理水井的补偿数据"""
    new_data = [
        ['水井', '眼', float(number), 500.00, shuijing],
    ]
    data_into_excel(sheet_name, new_data)

def handle_shuiguan(sheet_name, number, shuiguan):
    """处理给水管的补偿数据"""
    new_data = [
        ['给水管', 'm', float(number), 7.00, shuiguan],
    ]
    data_into_excel(sheet_name, new_data)

def handle_dijiao(sheet_name, number, dijiao):
    """处理地窖的补偿数据"""
    new_data = [
        ['地窖', '座', float(number), 800.00, dijiao],
    ]
    data_into_excel(sheet_name, new_data)

def handle_shuichi(data_list, sheet_name, area, year, roman, volume, yingxiang, buchang, yingxiangdanjia, buchangdanjia, shuichi):
    if '地上归' in data_list[1]:
        """处理地上附着物非户主所有的情况"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
        ]
        data_into_excel(sheet_name, new_data)

        new_data2 = [
            ['浆砌水池', 'm3', volume, 440.00, shuichi],
        ]
        name = data_list[1].split('归')[1].split('）')[0]
        return new_data2, name
    else:
        """处理浆砌水池的补偿数据"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
            ['浆砌水池', 'm3', volume, 440.00, shuichi],
        ]
        data_into_excel(sheet_name, new_data)
        return None, None  # 确保在所有情况下都有返回值

def handle_yutang(data_list, sheet_name, area, year, roman, volume, yingxiang, buchang,yingxiangdanjia, buchangdanjia, yutang, yumiao):
    if '地上归' in data_list[1]:
        """处理地上附着物非户主所有的情况"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
        ]
        data_into_excel(sheet_name, new_data)

        new_data2 = [
            ['土鱼塘', 'm3', volume, 7.40, yutang],
            ['鱼损', '亩', area, 1000.00, yumiao],
        ]
        name = data_list[1].split('归')[1].split('）')[0]
        return new_data2, name
    else:
        """处理土鱼塘的补偿数据"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
            ['土鱼塘', 'm3', volume, 7.40, yutang],
            ['鱼损', '亩', area, 1000.00, yumiao],
        ]
        data_into_excel(sheet_name, new_data)
        return None, None  # 确保在所有情况下都有返回值

def handle_zhaijidi(sheet_name, area, year, roman, buchang, yingxiang, buchangdanjia, yingxiangdanjia, lingxing):
    """处理宅基地的补偿数据"""
    new_data = [
        [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
        [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
        ['房前屋后零星林木', '户', 1.000, 1000.00, lingxing],
    ]
    data_into_excel(sheet_name, new_data)

def handle_default(data_list, sheet_name, tree_type, area, year, roman, yingxiang, buchang, yingxiangdanjia, buchangdanjia, tree, treedanjia):
    if '地上归' in data_list[1]:
        """处理地上附着物非户主所有的情况"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
        ]
        data_into_excel(sheet_name, new_data)
        tree_type = tree_type.split('（')[0]
        new_data2 = [
            [tree_type, '亩', area, treedanjia, tree],
        ]
        name = data_list[1].split('归')[1].split('）')[0]
        return new_data2, name
    else:
        """默认情况（其他类型）"""
        new_data = [
            [f'临时用地{year}年补偿', '亩', area, buchangdanjia, buchang],
            [f'临时用地影响期1年', '亩', area, yingxiangdanjia, yingxiang],
            [tree_type, '亩', area, treedanjia, tree],
        ]
        data_into_excel(sheet_name, new_data)
        return None, None  # 确保在所有情况下都有返回值


def village_jiti_into_excel(sheet_name, area, cunjiti, cunjitidanjia):
    new_data = [
        ['村集体土地补偿费', '亩', area, cunjitidanjia, cunjiti],
    ]
    data_into_excel(sheet_name, new_data)

def data_into_excel(sheet_name, new_data):
    """根据类型选择对应的函数，写入 Excel"""
    # 读取 Excel 文件
    wb = load_workbook('output_file.xlsx')
    ws = wb[sheet_name]
    # 设置列宽
    column_widths = {
        1: 20,
        2: 7,
        3: 12,
        4: 16,
        5: 16,
        6: 10
    }
    # 遍历指定的列和对应的宽度
    for col, width in column_widths.items():
        column_letter = get_column_letter(col)  # 获取列字母
        ws.column_dimensions[column_letter].width = width  # 设置列宽

    # 找到最后一行
    last_row = ws.max_row
    # 记录数据起始行
    start_row = last_row + 1
    # 将新数据写入工作表
    for row in new_data:
        last_row += 1
        for col, value in enumerate(row, start=1):
            unit = row[1] if isinstance(row[1], str) else ''
            # 只对数值做小数处理
            if isinstance(value, (int, float)):
                if unit in ('m2', 'm3'):
                    value = round_half_up(value, 2)
                    cell_format = '0.00'
                else:
                    value = round_half_up(value, 3)
                    cell_format = '0.000'
            elif col in [4, 5]:  # D、E列 (第4、5列) 保留2位小数
                value = round_half_up(value, 2) if isinstance(value, (int, float)) else value
                cell_format = '0.00'  # Excel 显示 2 位小数
            else:
                cell_format = None  # 其他列不做特殊格式化

            cell = ws.cell(row=last_row, column=col, value=value)

            # 设置 Excel 显示格式
            if cell_format:
                cell.number_format = cell_format  # 确保 Excel 显示小数
                cell.alignment = Alignment(horizontal='right', vertical='center')  # 水平左对齐，垂直居中
            else:
                # A 列左对齐，其他列居中
                if col == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')  # A列左对齐
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')  # 其他列居中
    # 保存 Excel
    wb.save('output_file.xlsx')
def other_people_into_excel(other_dict):
    for keys, values in other_dict.items():
        if keys == "基本信息":
            continue
        try:
            for value in values:
                data_into_excel(keys, value)
        except:
            village_name = other_dict['基本信息'][0]
            date = other_dict['基本信息'][1]
            excel_header = other_dict['基本信息'][2]
            header_into_excel(keys, village_name, date, excel_header)
            for value in values:
                data_into_excel(keys, value)

def summary_into_excel(sheet_name):
    # 加载 Excel 文件
    wb = load_workbook('output_file.xlsx')
    # 遍历所有 Sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 记录数据起始行
        start_row = 4
        # 记录最后一行
        last_row = ws.max_row
        # 计算 E 列总和
        sum_formula = round(sum(ws.cell(row=row, column=5).value for row in range(start_row, last_row + 1)),2)
        # 转换为中文大写
        total_chinese = cn2an.an2cn(sum_formula, "rmb")
        # 转换为 **繁体中文**
        total_chinese_traditional = HanziConv.toTraditional(total_chinese)
        # 在 E 列插入公式，并设置保留两位小数
        cell = ws.cell(row=last_row + 1, column=5, value=sum_formula)
        cell.number_format = '0.00'  # 保留2位小数

        # 设置单元格为垂直居中，左对齐
        cell.alignment = Alignment(horizontal='right', vertical='center')
        # 在 A 列最后一行填入 "合计"
        ws.cell(row=last_row + 1, column=1, value="合    计")
        ws.cell(row=last_row + 2, column=1, value="大    写")
        ws.cell(row=last_row + 2, column=2, value="人民币")
        merge_range = f"C{last_row + 2}:F{last_row + 2}"
        ws.merge_cells(merge_range)
        ws.cell(row=last_row + 2, column=3, value=total_chinese)
        # 设置居中对齐
        for row in [last_row + 1, last_row + 2]:
            for col in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # 保存 Excel
    wb.save('output_file.xlsx')
def sort_sheet():
    # 加载工作簿
    wb = load_workbook("output_file.xlsx")

    # 指定要移动的 Sheet 名称
    target_sheet_name = "村集体"
    try:
        # 找到该 Sheet 对象
        target_sheet = wb[target_sheet_name]

        # 从工作簿的 sheets 列表中移除，再添加到末尾
        wb._sheets.remove(target_sheet)
        wb._sheets.append(target_sheet)
    except:
        pass

    # 保存文件
    wb.save("output_file.xlsx")

if __name__ == '__main__':

    header_into_excel(name='qqqqqqqqq', village_name='sssss', date=datetime.now().strftime('%Y-%m-%d'), excel_header='sdawdsdads')