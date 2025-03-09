from datetime import datetime
from openpyxl.styles import Alignment
import pandas as pd
import os
from openpyxl import load_workbook

# Excel 文件名
file_name = 'new_file.xlsx'

# 获取当前应该使用的 sheet 编号
if os.path.exists(file_name):
    # 读取已有的 Excel 文件，获取所有 Sheet 名称
    wb = load_workbook(file_name)
    existing_sheets = wb.sheetnames

    # 找到现有最大编号，确保递增
    max_num = 1  # 如果找不到，则默认从 1 开始
    for sheet in existing_sheets:
        if sheet.startswith('NewSheet'):
            try:
                sheet_num = int(sheet.replace('NewSheet', ''))
                max_num = max(max_num, sheet_num)
            except ValueError:
                continue  # 如果 Sheet 名不是数字结尾，就跳过

    num = max_num + 1  # 下一个 Sheet 号
else:
    num = 1  # 如果文件不存在，则从 1 开始

# 动态生成列名
name = 'heng'
village_name = 'aaaa'
date = datetime.date
new_data = {
    '户号0000%d' % num: [],
    '户主：%s' % name: [],
    '住址：%s' % village_name: [],
    '%s' % date: []
}
df_new = pd.DataFrame(new_data)

# 在 'Column1' 和 'Column2' 之间插入一个空列
df_new.insert(1, '', '')  # 在索引位置 1 插入一个空列


# **写入 Excel**
if not os.path.exists(file_name):
    # **文件不存在，创建新的 Excel**
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df_new.to_excel(writer, sheet_name=f'NewSheet{num}', index=False)
else:
    # **文件已存在，追加 Sheet**
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df_new.to_excel(writer, sheet_name=f'NewSheet{num}', index=False)

print(f"数据已写入 {file_name} 的 NewSheet{num} 工作表。")
