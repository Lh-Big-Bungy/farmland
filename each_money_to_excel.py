from openpyxl import Workbook, load_workbook
from decimal import Decimal, ROUND_HALF_UP
from openpyxl.styles import Alignment, Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00  # 数字格式

def get_data():
    # 读取 Excel 文件
    file_path = "output_file.xlsx"  # 你的 Excel 文件
    wb = load_workbook(file_path)
    # 存储数据
    data_list = []
    # 获取所有 sheet 名称，并排除最后一个
    sheets = wb.sheetnames
    # 读取每个 sheet 的数据
    for sheet in sheets:
        ws = wb[sheet]
        # 获取 E 列的数据
        e_column = ws['E']  # 获取 E 列的所有单元格
        # 获取倒数第二行的数据
        money = e_column[-2].value  # 倒数第二行
        name = ws['B2'].value.split(":")[1]
        village = ws['D2'].value.split(':')[1]
        header = ws['A1'].value.split('兑付表')[0] + '公示表'
        # 将数据存储为元组
        data_tuple = (name, Decimal(str(money)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP))  # 转换为 Decimal 并保留两位小数
        data_list.append(data_tuple)
    print(header)
    print(village)
    return data_list, village, header
def each_money_to_excel(data_list, village, header):
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    # 合并 A3 到 A4 单元格
    ws.merge_cells('A1:E1')
    ws['A1'] = header
    # 设置 A1 单元格字体大小为 14 并加粗
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')  # 居中
    ws.insert_rows(2)  # 在第 2 行插入一行空白行
    ws['E2'] = "单位：元"
    ws['A3'] = "序号"
    ws['B3'] = "姓名"
    ws['C3'] = "村组"
    ws['D3'] = "货币安置\n补偿金额"
    ws['E3'] = "备注"
    # 设置表头样式
    for cell in ws[3]:
        # cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 填充数据
    total = Decimal('0.00')  # 合计金额，初始化为 Decimal 类型
    for index, data in enumerate(data_list, start=1):
        ws[f'A{index + 3}'] = index  # 序号
        ws[f'B{index + 3}'] = data[0]  # 姓名
        ws[f'C{index + 3}'] = village  # 村组
        ws[f'D{index + 3}'] = data[1]  # 补偿金额
        ws[f'D{index + 3}'].number_format = FORMAT_NUMBER_00  # 设置 D 列为两位小数

        total += data[1]  # 累加金额
    # 添加合计行
    last_row = len(data_list) + 4
    ws[f'A{last_row}'] = "合    计"
    ws[f'D{last_row}'] = total
    ws[f'D{last_row}'].number_format = FORMAT_NUMBER_00  # 设置合计行为两位小数

    # 设置合计行样式
    ws[f'A{last_row}'].font = Font(bold=True)
    ws[f'D{last_row}'].font = Font(bold=True)
    # 设置 A、B、C 列内容居中
    for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save("补偿公示表.xlsx")

def each_run():
    data_list, village, header = get_data()
    each_money_to_excel(data_list, village, header)
if __name__ == '__main__':
    each_run()