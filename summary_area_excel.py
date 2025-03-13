from openpyxl import Workbook, load_workbook
from decimal import Decimal
from openpyxl.styles import Alignment, Border, Side, Font


def get_summary_money():
    # 读取 Excel 文件
    file_path = "output_file.xlsx"  # 你的 Excel 文件
    wb = load_workbook(file_path)

    # 获取所有 sheet 名称，并排除最后一个
    sheets = wb.sheetnames

    # 存储数据
    data_dict = {}

    # 读取每个 sheet 的数据
    for sheet in sheets:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取，跳过表头
            if len(row) < 5:
                continue  # 确保数据完整

            project, unit, quantity, price, subtotal = row[:5]  # 取前 5 列
            if not isinstance(quantity, (int, float)):
                continue  # 跳过无效数据
            # 使用 Decimal 来保证精度
            area = Decimal(str(quantity))
            key = project  # 以 (项目, 单位) 作为唯一标识
            data_dict[key] = data_dict.get(key, Decimal('0')) + area  # 累加面积  防止精度错误
    print(data_dict)
    ws = wb.active
    village_name = ws['D2'].value.split(':')[1]
    header = ws['A1'].value.split('兑付表')[0] + '分类面积汇总表'
    key_list = []
    print(village_name, header)
    for key in data_dict.keys():
        if key not in key_list:
            key_list.append(key)
    # 要删除的项
    items_to_remove = ['土地补偿费（户）', '土地安置补助费', '耕地青苗费', '村集体土地补偿费', '合计', '小计']
    for i in items_to_remove:
        if i in key_list:
            key_list.remove(i)
    fenmu_list = []
    for j in key_list[:]:  # 修改列表时同时遍历列表，可能会导致跳过某些项，使用副本进行遍历
        if '坟' in j:
            fenmu_list.append(j)
            key_list.remove(j)
    print(key_list)
    print(fenmu_list)
    print(data_dict)
    return data_dict, village_name, header, key_list, fenmu_list
def summary_area_excel(header, village, data_dict, key_list, fenmu_list):
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    length = 0  # 定义
    fenmu_length = 0
    # 合并 A3 到 A4 单元格
    ws.merge_cells('A3:A4')
    ws['A3'] = '乡镇村'
    ws['A5'] = village
    # 合并 B2 到 B4 单元格
    ws.merge_cells('B3:D3')
    ws['B3'] = '土地补偿面积'
    ws['B4'] = '分户土地补偿面积'
    ws['B5'] = str(data_dict['土地补偿费（户）']) + '亩'
    ws['C4'] = '村集体土地补偿面积'
    ws['C5'] = str(data_dict['村集体土地补偿费']) + '亩'
    ws['D4'] = '小计'
    ws['D5'] = str(Decimal(str(data_dict['土地补偿费（户）'])) + Decimal(str(data_dict['村集体土地补偿费']))) + '亩'  # 防止精度错误
    if '耕地青苗费' in data_dict:
        # 合并 E3 到 E4 单元格
        ws.merge_cells('E3:E4')
        ws['E3'] = '土地安置面积'
        ws['E5'] = str(data_dict['土地安置补助费']) + '亩'
        # 合并 E3 到 E4 单元格
        ws.merge_cells('F3:F4')
        ws['F3'] = '耕地青苗面积'
        ws['F5'] = str(data_dict['耕地青苗费']) + '亩'
    else:
        # 合并 E3 到 E4 单元格
        ws.merge_cells('E3:F4')
        ws['E3'] = '土地安置面积'
        ws['E4'] = str(data_dict['土地安置补助费'] + '亩')
    if key_list:
        # 计算小计并填入最后一个单元格下面
        length = len(key_list)
        # 动态生成合并区域，假设从 G3 开始
        merge_range = f'G3:{chr(65 + 6 + length - 1)}3'  # chr(65) 是字母 'A' 的 ASCII 值，65 + 6 是G， 再加length然后-1 会给出正确的字母

        # 合并单元格
        ws.merge_cells(merge_range)
        ws['G3'] = '地上附着物'
        other_dict = {
            '浆砌水池': 'm3',
            '土鱼塘': 'm3',
            '鱼损': '亩',
            '晒场硬化': 'm2',
            '蔬菜大棚拆迁（钢结构）': 'm2',
            '水井': '眼',
            '给水管': 'm',
            '地窖': '座',
        }
        # 从 G4 开始横向循环填充数据
        for i in range(length):
            col = chr(65 + 6 + i)  # 从 G 列开始 (G -> 7, H -> 8, I -> 9 ...)
            if key_list[i] in ['浆砌水池', '土鱼塘', '鱼损', '晒场硬化', '蔬菜大棚拆迁（钢结构）', '水井', '给水管', '地窖']:
                ws[f'{col}4'] = key_list[i]
                ws[f'{col}5'] = str(data_dict[key_list[i]]) + other_dict[key_list[i]]  # 填入数据
            else:
                ws[f'{col}4'] = key_list[i]
                ws[f'{col}5'] = str(data_dict[key_list[i]]) + '亩'
    if fenmu_list:
        # 计算小计并填入最后一个单元格下面
        total_fenmu = sum(data_dict[fenmu] for fenmu in fenmu_list)
        fenmu_list.append('小计')
        fenmu_length = len(fenmu_list)
        # 确定下一个起始列，即从 G3 合并结束的列之后
        start_col = chr(65 + 6 + length)  # 计算从 G3 合并区域后开始的位置
        # 动态生成合并区域，假设从 start_col 和 G6 开始
        merge_range = f'{start_col}3:{chr(65 + 6 + length + fenmu_length - 1)}3'  # 合并区域从下一个位置开始
        # 合并单元格
        ws.merge_cells(merge_range)
        ws[start_col + '3'] = '坟墓'
        for i in range(fenmu_length):
            col = chr(65 + 6 + length + i)  # 从 G 列合并后的位置继续
            if i != fenmu_length - 1:
                ws[f'{col}4'] = fenmu_list[i]
                ws[f'{col}5'] = str(data_dict[fenmu_list[i]]) + '座'
            else:
                ws[f'{col}4'] = fenmu_list[i]
                ws[f'{col}5'] = str(total_fenmu) + '座'
    max_length = length + fenmu_length
    last_col = chr(65 + 6 + max_length - 1)
    # 合并 A1 到 X1 单元格
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = header
    # 插入单位
    # ws[f'{last_col}2'] = '单位：亩'
    # 设置所有列宽为 10
    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column  # 获取列索引
        ws.column_dimensions[chr(64 + col_letter)].width = 12  # 64 + 索引 转换成字母

    # 定义边框样式（细边框）
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # 遍历所有单元格，设置居中
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = border_style  # 添加边框
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")  # 居中

    # 设置 A1 单元格字体大小为 14 并加粗
    ws['A1'].font = Font(size=14, bold=True)
    # 保存Excel文件
    wb.save("补偿面积分类汇总表.xlsx")

if __name__ == '__main__':
    data_dict, village_name, header, key_list, fenmu_list = get_summary_money()
    summary_area_excel(header, village_name, data_dict, key_list, fenmu_list)